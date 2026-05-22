"""
物料管理 API 路由
功能：物料信息的增删改查
"""
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse, Response
from sqlalchemy.orm import Session
from sqlalchemy import func, or_
from typing import Optional
from datetime import datetime
import csv
import io
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from app.database import get_db
from app.models.material import Material
from app.models.event_data import EventData
from app.schemas.material import (
    MaterialQuery, MaterialList, MaterialDetail, MaterialCreate, MaterialUpdate,
    MATERIAL_TYPE_CHOICES
)

router = APIRouter()


# ============ 固定路径路由（按顺序定义） ============

@router.get("/", response_model=MaterialList)
async def list_materials(
    query: MaterialQuery = Depends(),
    db: Session = Depends(get_db)
):
    """
    查询物料列表（支持多条件组合查询）

    查询条件包括：
    - U9 物料号：u9_material_code（支持模糊查询）
    - 零件号：part_number（支持模糊查询）
    - 物料名称：product_name（支持模糊查询）
    - 物料分类：category
    - 项目：project（支持模糊查询）
    - 物料类型：material_type
    - 状态：status
    """
    db_query = db.query(Material)

    # 关键词搜索（U9物料号 或 物料名称，OR 逻辑）
    if query.keyword:
        db_query = db_query.filter(
            or_(
                Material.u9_material_code.contains(query.keyword),
                Material.product_name.contains(query.keyword),
                Material.part_number.contains(query.keyword),
            )
        )

    # U9 物料号筛选
    if query.u9_material_code:
        db_query = db_query.filter(Material.u9_material_code.contains(query.u9_material_code))

    # 零件号筛选
    if query.part_number:
        db_query = db_query.filter(Material.part_number.contains(query.part_number))

    # 物料名称筛选
    if query.product_name:
        db_query = db_query.filter(Material.product_name.contains(query.product_name))

    # 物料分类筛选
    if query.category:
        db_query = db_query.filter(Material.category == query.category)

    # 项目筛选
    if query.project:
        db_query = db_query.filter(Material.project.contains(query.project))

    # 车间筛选
    if query.workshop:
        db_query = db_query.filter(Material.workshop.contains(query.workshop))

    # 物料类型筛选
    if query.material_type:
        db_query = db_query.filter(Material.material_type == query.material_type)

    # 状态筛选
    if query.status:
        db_query = db_query.filter(Material.status == query.status)

    # 获取总数
    total = db_query.count()

    # 分页
    offset = (query.page - 1) * query.page_size
    items = db_query.order_by(Material.created_at.desc()).offset(offset).limit(query.page_size).all()

    return MaterialList(total=total, items=items)


@router.get("/download-template")
async def download_template():
    """下载物料导入模板（xlsx 格式）"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "物料导入模板"

    header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_alignment = Alignment(vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    headers = [
        'U9 物料号 *',
        '零件号',
        '物料名称 *',
        '物料描述',
        '规格型号',
        '分类',
        '项目',
        '单位',
        '单件工时',
        '物料类型',
        '状态'
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 25
    ws.column_dimensions['H'].width = 10
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 12

    ws.row_dimensions[1].height = 30

    example_data = [
        ['U9-2024-001', 'PN-001', '示例物料1', '这是一个示例物料', '规格A', '电子元件', '项目A', '个', 0.5, 'product', 'active'],
        ['U9-2024-002', 'PN-002', '示例物料2', '这是另一个示例物料', '规格B', '机械零件', '项目B', '件', 1.0, 'material', 'active'],
    ]

    for row_num, row_data in enumerate(example_data, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border

    note_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    note_font = Font(name='微软雅黑', size=10, color='FF6600')

    ws.merge_cells('A11:K11')
    note_cell = ws.cell(row=11, column=1, value='说明：带 * 号为必填字段；物料类型可填 product(产品)/semi_finished(半成品)/material(原材料)/auxiliary(辅料)；状态只能填写 active 或 inactive')
    note_cell.font = note_font
    note_cell.fill = note_fill

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        content=output.getvalue(),
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            'Content-Disposition': "attachment; filename*=UTF-8''%E7%89%A9%E6%96%99%E5%AF%BC%E5%85%A5%E6%A8%A1%E6%9D%BF.xlsx"
        }
    )


@router.post("/import")
async def import_materials(
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """从 CSV 或 XLSX 文件导入物料数据"""
    if not (file.filename.endswith('.csv') or file.filename.endswith('.xlsx')):
        raise HTTPException(status_code=400, detail="仅支持 CSV 或 XLSX 格式文件")

    try:
        content = await file.read()

        rows_data = []

        def normalize_header(h):
            """去除表头中的 * 号和空白"""
            if h is None:
                return ''
            return h.replace('*', '').strip()

        if file.filename.endswith('.csv'):
            # 处理 CSV 文件
            encodings_to_try = ['utf-8', 'gbk', 'gb2312', 'gb18030', 'utf-8-sig']
            csv_content = None

            for encoding in encodings_to_try:
                try:
                    csv_content = content.decode(encoding)
                    break
                except UnicodeDecodeError:
                    continue

            if csv_content is None:
                raise HTTPException(status_code=400, detail="无法识别的文件编码，请确保 CSV 文件使用 UTF-8 或 GBK 编码")

            reader = csv.DictReader(io.StringIO(csv_content))
            for row in reader:
                normalized = {normalize_header(k): v for k, v in row.items()}
                rows_data.append(normalized)
        else:
            # 处理 XLSX 文件
            wb = openpyxl.load_workbook(io.BytesIO(content))
            ws = wb.active

            # 读取表头
            raw_headers = [cell.value for cell in ws[1]]
            headers = [normalize_header(h) for h in raw_headers]

            # 读取数据行
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                row_dict = {}
                for col_idx, value in enumerate(row):
                    if col_idx < len(headers):
                        row_dict[headers[col_idx]] = str(value) if value is not None else ''
                rows_data.append(row_dict)

        imported_count = 0
        updated_count = 0
        errors = []

        for row_num, row in enumerate(rows_data, start=2):
            try:
                # 检查必填字段
                u9_material_code = row.get('U9 物料号', '').strip()
                material_name = row.get('物料名称', '').strip()

                if not u9_material_code or not material_name:
                    errors.append(f"第{row_num}行：缺少必填字段")
                    continue

                part_number = row.get('零件号', '').strip()

                # 物料类型
                material_type = row.get('物料类型', 'product').strip()
                if material_type not in MATERIAL_TYPE_CHOICES:
                    material_type = 'product'

                # 检查 U9 物料号是否已存在
                existing = db.query(Material).filter(
                    Material.u9_material_code == u9_material_code
                ).first()

                if existing:
                    # 更新现有物料
                    existing.part_number = part_number
                    existing.product_name = material_name
                    existing.description = row.get('物料描述', '').strip() or None
                    existing.specification = row.get('规格型号', '').strip() or None
                    existing.category = row.get('分类', '').strip() or None
                    existing.project = row.get('项目', '').strip() or None
                    existing.unit = row.get('单位', '').strip() or None
                    unit_work_time_str = row.get('单件工时', '').strip()
                    existing.unit_work_time = float(unit_work_time_str) if unit_work_time_str else None
                    existing.material_type = material_type
                    status = row.get('状态', 'active').strip()
                    existing.status = status if status in ['active', 'inactive'] else 'active'
                    existing.updated_at = datetime.now()
                    updated_count += 1
                else:
                    # 创建新物料
                    status = row.get('状态', 'active').strip()
                    unit_work_time_str = row.get('单件工时', '').strip()
                    unit_work_time = float(unit_work_time_str) if unit_work_time_str else None
                    material = Material(
                        u9_material_code=u9_material_code,
                        part_number=part_number,
                        product_name=material_name,
                        description=row.get('物料描述', '').strip() or None,
                        specification=row.get('规格型号', '').strip() or None,
                        category=row.get('分类', '').strip() or None,
                        project=row.get('项目', '').strip() or None,
                        unit=row.get('单位', '').strip() or None,
                        unit_work_time=unit_work_time,
                        material_type=material_type,
                        status=status if status in ['active', 'inactive'] else 'active'
                    )
                    db.add(material)
                    imported_count += 1

            except Exception as e:
                errors.append(f"第{row_num}行：{str(e)}")

        # 提交事务
        db.commit()

        return {
            "message": "导入完成",
            "imported": imported_count,
            "updated": updated_count,
            "errors": errors[:10]  # 最多返回 10 条错误信息
        }

    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"导入失败：{str(e)}")


@router.get("/export")
async def export_materials(
    u9_material_code: Optional[str] = None,
    part_number: Optional[str] = None,
    product_name: Optional[str] = None,
    category: Optional[str] = None,
    project: Optional[str] = None,
    material_type: Optional[str] = None,
    status: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """导出物料数据为 CSV 文件"""
    db_query = db.query(Material)

    # 应用筛选条件
    if u9_material_code:
        db_query = db_query.filter(Material.u9_material_code.contains(u9_material_code))
    if part_number:
        db_query = db_query.filter(Material.part_number.contains(part_number))
    if product_name:
        db_query = db_query.filter(Material.product_name.contains(product_name))
    if category:
        db_query = db_query.filter(Material.category == category)
    if project:
        db_query = db_query.filter(Material.project.contains(project))
    if material_type:
        db_query = db_query.filter(Material.material_type == material_type)
    if status:
        db_query = db_query.filter(Material.status == status)

    materials = db_query.order_by(Material.created_at.desc()).all()

    # 创建 CSV 文件
    output = io.StringIO()
    writer = csv.writer(output)

    # 写入表头
    writer.writerow([
        'ID', 'U9 物料号', '零件号', '物料名称', '物料描述',
        '规格型号', '分类', '项目', '单位', '单件工时', '物料类型', '状态', '创建时间', '更新时间'
    ])

    # 写入数据
    for material in materials:
        writer.writerow([
            material.id,
            material.u9_material_code,
            material.part_number,
            material.product_name,
            material.description or '',
            material.specification or '',
            material.category or '',
            material.project or '',
            material.unit or '',
            material.unit_work_time or '',
            MATERIAL_TYPE_CHOICES.get(material.material_type, material.material_type),
            material.status,
            material.created_at.strftime('%Y-%m-%d %H:%M:%S') if material.created_at else '',
            material.updated_at.strftime('%Y-%m-%d %H:%M:%S') if material.updated_at else ''
        ])

    # 生成文件
    csv_content = output.getvalue()
    filename = f"materials_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

    return StreamingResponse(
        iter([csv_content]),
        media_type='text/csv; charset=utf-8',
        headers={
            'Content-Disposition': f"attachment; filename*=UTF-8''{filename}"
        }
    )


@router.get("/stats/summary")
async def get_material_stats(db: Session = Depends(get_db)):
    """获取物料统计信息"""
    total = db.query(func.count(Material.id)).scalar()
    active = db.query(func.count(Material.id)).filter(Material.status == 'active').scalar()
    inactive = db.query(func.count(Material.id)).filter(Material.status == 'inactive').scalar()

    # 按分类统计
    category_stats = db.query(
        Material.category,
        func.count(Material.id).label('count')
    ).filter(
        Material.category.isnot(None)
    ).group_by(Material.category).all()

    # 按物料类型统计
    type_stats = db.query(
        Material.material_type,
        func.count(Material.id).label('count')
    ).group_by(Material.material_type).all()

    return {
        "total": total,
        "active": active,
        "inactive": inactive,
        "categories": [{"category": cat, "count": count} for cat, count in category_stats],
        "material_types": [{"type": t, "count": count} for t, count in type_stats]
    }


# ============ 动态路径路由（放在最后） ============

@router.get("/{material_id}", response_model=MaterialDetail)
async def get_material(material_id: int, db: Session = Depends(get_db)):
    """获取物料详情"""
    material = db.query(Material).filter(Material.id == material_id).first()
    if not material:
        raise HTTPException(status_code=404, detail="物料不存在")
    return material


@router.post("/", response_model=MaterialDetail)
async def create_material(material: MaterialCreate, db: Session = Depends(get_db)):
    """创建新物料"""
    # 检查 U9 物料号是否已存在
    existing = db.query(Material).filter(Material.u9_material_code == material.u9_material_code).first()
    if existing:
        raise HTTPException(status_code=400, detail="U9 物料号已存在")

    db_material = Material(**material.model_dump())
    db.add(db_material)
    db.commit()
    db.refresh(db_material)
    return db_material


@router.put("/{material_id}", response_model=MaterialDetail)
async def update_material(material_id: int, material: MaterialUpdate, db: Session = Depends(get_db)):
    """更新物料信息"""
    db_material = db.query(Material).filter(Material.id == material_id).first()
    if not db_material:
        raise HTTPException(status_code=404, detail="物料不存在")

    # 检查 U9 物料号是否与其他物料冲突
    if material.u9_material_code:
        existing = db.query(Material).filter(
            Material.u9_material_code == material.u9_material_code,
            Material.id != material_id
        ).first()
        if existing:
            raise HTTPException(status_code=400, detail="U9 物料号已存在")

    # 更新字段
    update_data = material.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(db_material, field, value)

    db_material.updated_at = datetime.now()
    db.commit()
    db.refresh(db_material)
    return db_material


@router.delete("/{material_id}")
async def delete_material(material_id: int, db: Session = Depends(get_db)):
    """删除物料"""
    material = db.query(Material).filter(Material.id == material_id).first()
    if not material:
        raise HTTPException(status_code=404, detail="物料不存在")

    db.delete(material)
    db.commit()
    return {"message": "物料已删除"}


@router.get("/{material_id}/events")
async def get_material_events(
    material_id: int,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
):
    """获取物料关联的加工事件（通过 u9_material_code 或 part_number 匹配 start_code）"""
    material = db.query(Material).filter(Material.id == material_id).first()
    if not material:
        raise HTTPException(status_code=404, detail="物料不存在")

    # 通过 u9_material_code 和 part_number 匹配 event_data.start_code
    conditions = [EventData.start_code == material.u9_material_code]
    if material.part_number and material.part_number != material.u9_material_code:
        conditions.append(EventData.start_code == material.part_number)

    query = db.query(EventData).filter(or_(*conditions))

    total = query.count()
    items = query.order_by(EventData.start_time.desc()).offset(
        (page - 1) * page_size
    ).limit(page_size).all()

    return {
        "total": total,
        "items": [
            {
                "id": e.id,
                "event_uid": e.event_uid,
                "start_code": e.start_code,
                "start_time": e.start_time,
                "end_time": e.end_time,
                "duringtime": e.duringtime,
                "machine_id": e.machine_id,
                "operator_name": e.operator_name,
                "process_no": e.process_no,
                "line_code": e.line_code,
                "created_at": str(e.created_at) if e.created_at else None,
            }
            for e in items
        ],
        "material": {
            "id": material.id,
            "u9_material_code": material.u9_material_code,
            "part_number": material.part_number,
            "product_name": material.product_name,
        }
    }
