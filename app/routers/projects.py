"""
项目管理 API 路由
"""
from fastapi import APIRouter, Depends, HTTPException, Query, File, UploadFile
from fastapi.responses import Response, StreamingResponse
from sqlalchemy.orm import Session
from typing import Optional
from datetime import datetime
import logging
import csv
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logger = logging.getLogger(__name__)

from app.database import get_db
from app.models.project import Project
from app.schemas.project import (
    ProjectCreate, ProjectUpdate, ProjectDetail, ProjectList
)

router = APIRouter()


@router.get("/", response_model=ProjectList)
async def list_projects(
    name: Optional[str] = Query(default=None, description="项目名称"),
    code: Optional[str] = Query(default=None, description="项目编码"),
    customer: Optional[str] = Query(default=None, description="客户"),
    status: Optional[str] = Query(default=None, description="状态"),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=20, ge=1, le=100),
    db: Session = Depends(get_db)
):
    """查询项目列表"""
    query = db.query(Project)

    if name:
        query = query.filter(Project.name.contains(name))
    if code:
        query = query.filter(Project.code.contains(code))
    if customer:
        query = query.filter(Project.customer.contains(customer))
    if status:
        query = query.filter(Project.status == status)

    total = query.count()
    items = query.order_by(Project.sort_order, Project.id).offset(
        (page - 1) * page_size
    ).limit(page_size).all()

    return ProjectList(total=total, items=items)


@router.get("/all")
async def list_all_projects(db: Session = Depends(get_db)):
    """获取所有启用的项目（用于下拉选择）"""
    projects = db.query(Project).filter(
        Project.status == 'active'
    ).order_by(Project.sort_order, Project.id).all()
    return {"items": projects}


STATUS_CHOICES = {'active', 'completed', 'suspended', 'cancelled'}


@router.get("/download-template")
async def download_template():
    """下载项目导入模板（xlsx 格式）"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "项目导入模板"

    header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_alignment = Alignment(vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    headers = ['项目编码 *', '项目名称 *', '项目描述', '客户', '负责人', '开始日期', '结束日期', '状态', '排序']

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    widths = [15, 30, 40, 20, 15, 15, 15, 12, 10]
    for i, w in enumerate(widths):
        ws.column_dimensions[chr(65 + i)].width = w

    ws.row_dimensions[1].height = 30

    example_data = [
        ['PRJ-001', '示例项目1', '这是一个示例项目', '客户A', '张三', '2024-01-01', '2024-12-31', 'active', 0],
        ['PRJ-002', '示例项目2', '另一个示例项目', '客户B', '李四', '2024-03-01', '2024-09-30', 'active', 1],
    ]

    for row_num, row_data in enumerate(example_data, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border

    note_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    note_font = Font(name='微软雅黑', size=10, color='FF6600')

    ws.merge_cells('A11:I11')
    note_cell = ws.cell(row=11, column=1, value='说明：带 * 号为必填字段；状态可填 active(进行中)/completed(已完成)/suspended(已暂停)/cancelled(已取消)')
    note_cell.font = note_font
    note_cell.fill = note_fill

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        content=output.getvalue(),
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            'Content-Disposition': "attachment; filename*=UTF-8''%E9%A1%B9%E7%9B%AE%E5%AF%BC%E5%85%A5%E6%A8%A1%E6%9D%BF.xlsx"
        }
    )


@router.post("/import")
async def import_projects(
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """从 CSV 或 XLSX 文件导入项目数据"""
    if not (file.filename.endswith('.csv') or file.filename.endswith('.xlsx')):
        raise HTTPException(status_code=400, detail="仅支持 CSV 或 XLSX 格式文件")

    try:
        content = await file.read()
        rows_data = []

        def normalize_header(h):
            """去除表头中的 * 号和空白"""
            if h is None:
                return ''
            return h.replace('*', '').strip()

        if file.filename.endswith('.csv'):
            encodings_to_try = ['utf-8', 'gbk', 'gb2312', 'gb18030', 'utf-8-sig']
            csv_content = None
            for encoding in encodings_to_try:
                try:
                    csv_content = content.decode(encoding)
                    break
                except UnicodeDecodeError:
                    continue
            if csv_content is None:
                raise HTTPException(status_code=400, detail="无法识别的文件编码，请确保 CSV 文件使用 UTF-8 或 GBK 编码")
            reader = csv.DictReader(io.StringIO(csv_content))
            for row in reader:
                normalized = {normalize_header(k): v for k, v in row.items()}
                rows_data.append(normalized)
        else:
            wb = openpyxl.load_workbook(io.BytesIO(content))
            ws = wb.active
            raw_headers = [cell.value for cell in ws[1]]
            headers = [normalize_header(h) for h in raw_headers]
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = {}
                for col_idx, value in enumerate(row):
                    if col_idx < len(headers):
                        row_dict[headers[col_idx]] = str(value) if value is not None else ''
                rows_data.append(row_dict)

        imported_count = 0
        updated_count = 0
        errors = []

        for row_num, row in enumerate(rows_data, start=2):
            try:
                code = row.get('项目编码', '').strip()
                name = row.get('项目名称', '').strip()

                if not code or not name:
                    errors.append(f"第{row_num}行：缺少必填字段（项目编码或项目名称）")
                    continue

                status = row.get('状态', 'active').strip()
                if status not in STATUS_CHOICES:
                    status = 'active'

                sort_order_str = row.get('排序', '0').strip()
                try:
                    sort_order = int(sort_order_str) if sort_order_str else 0
                except ValueError:
                    sort_order = 0

                existing = db.query(Project).filter(Project.code == code).first()

                if existing:
                    existing.name = name
                    existing.description = row.get('项目描述', '').strip() or None
                    existing.customer = row.get('客户', '').strip() or None
                    existing.manager = row.get('负责人', '').strip() or None
                    existing.start_date = row.get('开始日期', '').strip() or None
                    existing.end_date = row.get('结束日期', '').strip() or None
                    existing.status = status
                    existing.sort_order = sort_order
                    existing.updated_at = datetime.now()
                    updated_count += 1
                else:
                    project = Project(
                        code=code,
                        name=name,
                        description=row.get('项目描述', '').strip() or None,
                        customer=row.get('客户', '').strip() or None,
                        manager=row.get('负责人', '').strip() or None,
                        start_date=row.get('开始日期', '').strip() or None,
                        end_date=row.get('结束日期', '').strip() or None,
                        status=status,
                        sort_order=sort_order,
                    )
                    db.add(project)
                    imported_count += 1
            except Exception as e:
                errors.append(f"第{row_num}行：{str(e)}")

        db.commit()

        return {
            "message": "导入完成",
            "imported": imported_count,
            "updated": updated_count,
            "errors": errors[:10]
        }

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"导入失败：{str(e)}")


@router.get("/export")
async def export_projects(
    name: Optional[str] = None,
    code: Optional[str] = None,
    customer: Optional[str] = None,
    status: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """导出项目数据为 CSV 文件"""
    query = db.query(Project)

    if name:
        query = query.filter(Project.name.contains(name))
    if code:
        query = query.filter(Project.code.contains(code))
    if customer:
        query = query.filter(Project.customer.contains(customer))
    if status:
        query = query.filter(Project.status == status)

    projects = query.order_by(Project.sort_order, Project.id).all()

    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow(['ID', '项目编码', '项目名称', '项目描述', '客户', '负责人', '开始日期', '结束日期', '状态', '排序', '创建时间', '更新时间'])

    status_label_map = {'active': '进行中', 'completed': '已完成', 'suspended': '已暂停', 'cancelled': '已取消'}

    for p in projects:
        writer.writerow([
            p.id,
            p.code,
            p.name,
            p.description or '',
            p.customer or '',
            p.manager or '',
            p.start_date or '',
            p.end_date or '',
            status_label_map.get(p.status, p.status),
            p.sort_order or 0,
            p.created_at.strftime('%Y-%m-%d %H:%M:%S') if p.created_at else '',
            p.updated_at.strftime('%Y-%m-%d %H:%M:%S') if p.updated_at else '',
        ])

    csv_content = output.getvalue()
    filename = f"projects_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

    return StreamingResponse(
        iter([csv_content]),
        media_type='text/csv; charset=utf-8',
        headers={
            'Content-Disposition': f"attachment; filename*=UTF-8''{filename}"
        }
    )


@router.get("/{project_id}", response_model=ProjectDetail)
async def get_project(project_id: int, db: Session = Depends(get_db)):
    """获取项目详情"""
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="项目不存在")
    return project


@router.post("/", response_model=ProjectDetail)
async def create_project(project: ProjectCreate, db: Session = Depends(get_db)):
    """创建项目"""
    existing = db.query(Project).filter(
        (Project.name == project.name) | (Project.code == project.code)
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail="项目名称或编码已存在")

    db_project = Project(**project.model_dump())
    db.add(db_project)
    db.commit()
    db.refresh(db_project)
    return db_project


@router.put("/{project_id}", response_model=ProjectDetail)
async def update_project(project_id: int, project: ProjectUpdate, db: Session = Depends(get_db)):
    """更新项目"""
    db_project = db.query(Project).filter(Project.id == project_id).first()
    if not db_project:
        raise HTTPException(status_code=404, detail="项目不存在")

    if project.name:
        existing = db.query(Project).filter(
            Project.name == project.name, Project.id != project_id
        ).first()
        if existing:
            raise HTTPException(status_code=400, detail="项目名称已存在")

    if project.code:
        existing = db.query(Project).filter(
            Project.code == project.code, Project.id != project_id
        ).first()
        if existing:
            raise HTTPException(status_code=400, detail="项目编码已存在")

    update_data = project.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(db_project, field, value)

    db.commit()
    db.refresh(db_project)
    return db_project


@router.delete("/{project_id}")
async def delete_project(project_id: int, db: Session = Depends(get_db)):
    """删除项目"""
    db_project = db.query(Project).filter(Project.id == project_id).first()
    if not db_project:
        raise HTTPException(status_code=404, detail="项目不存在")

    db.delete(db_project)
    db.commit()
    return {"success": True, "message": "项目已删除"}
