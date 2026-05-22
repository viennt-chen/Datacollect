"""
设备管理 API 路由
功能：设备信息的增删改查，以及 ALARM、PV、SV、Event 数据查询
"""
from fastapi import APIRouter, Depends, HTTPException, Query, File, UploadFile
from fastapi.responses import Response, StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, and_, or_, desc
from typing import Optional, List, Dict, Any
from datetime import datetime, timedelta
import json
import zlib
import gzip
import struct
import re
import csv
import io
import urllib.parse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from app.database import get_db
from app.models.device import Device
from app.models.alarm_record import AlarmRecord
from app.models.pv_compressed_param import PVCompressedParam
from app.models.sv_compressed_param import SVCompressedParam
from app.models.event_data import EventData
from app.models.mqtt_topic_config import MQTTTopicConfig
from app.models.product_order import ProductOrder
from app.models.material import Material
from app.models.device_status_monitor_config import DeviceStatusMonitorConfig
from app.models.db_param_curve import DBParamCurve
from app.schemas.device import (
    DeviceQuery, DeviceList, DeviceDetail, DeviceCreate, DeviceUpdate
)
from app.services.device_data_collector import get_device_data_collector
from app.services.mqtt_collector import get_collector
from app.utils.rule_evaluator import (
    get_nested_value as _get_nested_value,
    match_value as _match_value,
    apply_extraction as _apply_extraction,
    evaluate_single_condition as _evaluate_single_condition,
    evaluate_config as _evaluate_config,
    load_device_rules,
    match_curve,
    evaluate_curve_config,
    evaluate_curve_config_from_buffer,
)

router = APIRouter(prefix="/devices", tags=["设备管理"])


def decompress_payload(compressed_data: bytes) -> Dict[str, Any]:
    """解压缩 payload 数据"""
    try:
        decompressed = gzip.decompress(compressed_data).decode('utf-8')
        return json.loads(decompressed)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"解压缩失败：{str(e)}")


def parse_device_mqtt_topics(device: Device) -> dict:
    """解析设备的 mqtt_topics 字段"""
    device_dict = {
        c.name: getattr(device, c.name) for c in device.__table__.columns
    }
    if device_dict.get('mqtt_topics'):
        try:
            device_dict['mqtt_topics'] = json.loads(device_dict['mqtt_topics'])
        except (json.JSONDecodeError, TypeError):
            device_dict['mqtt_topics'] = []
    else:
        device_dict['mqtt_topics'] = []
    return device_dict


@router.get("/", response_model=DeviceList)
async def list_devices(
    query: DeviceQuery = Depends(),
    db: Session = Depends(get_db)
):
    """
    查询设备列表（支持多条件组合查询）
    
    查询条件包括：
    - 设备编号：device_code（支持模糊查询）
    - 设备名称：device_name（支持模糊查询）
    - 设备类型：device_type
    - 状态：status
    - 产线：line_code
    - 工厂：factory_code
    """
    db_query = db.query(Device)
    
    # 设备编号筛选
    if query.device_code:
        db_query = db_query.filter(Device.device_code.contains(query.device_code))
    
    # 设备名称筛选
    if query.device_name:
        db_query = db_query.filter(Device.device_name.contains(query.device_name))
    
    # 设备类型筛选
    if query.device_type:
        db_query = db_query.filter(Device.device_type == query.device_type)
    
    # 状态筛选
    if query.status:
        db_query = db_query.filter(Device.status == query.status)
    
    # 产线筛选
    if query.line_code:
        db_query = db_query.filter(Device.line_code == query.line_code)
    
    # 工厂筛选
    if query.factory_code:
        db_query = db_query.filter(Device.factory_code == query.factory_code)
    
    # 获取总数
    total = db_query.count()
    
    # 分页
    offset = (query.page - 1) * query.page_size
    items = db_query.order_by(Device.created_at.desc()).offset(offset).limit(query.page_size).all()
    
    # 解析 mqtt_topics 字段
    parsed_items = []
    for item in items:
        item_dict = parse_device_mqtt_topics(item)
        parsed_items.append(DeviceDetail(**item_dict))
    
    return DeviceList(total=total, items=parsed_items)


@router.get("/stats/summary")
async def get_device_stats(db: Session = Depends(get_db)):
    """获取设备统计信息"""
    total = db.query(func.count(Device.device_code)).scalar()
    active = db.query(func.count(Device.device_code)).filter(Device.status == 'active').scalar()
    inactive = db.query(func.count(Device.device_code)).filter(Device.status == 'inactive').scalar()
    maintenance = db.query(func.count(Device.device_code)).filter(Device.status == 'maintenance').scalar()
    enabled = db.query(func.count(Device.device_code)).filter(Device.is_enabled == True).scalar()
    
    # 按设备类型统计
    type_stats = db.query(
        Device.device_type,
        func.count(Device.device_code).label('count')
    ).filter(
        Device.device_type.isnot(None)
    ).group_by(Device.device_type).all()
    
    # 按产线统计
    line_stats = db.query(
        Device.line_code,
        func.count(Device.device_code).label('count')
    ).filter(
        Device.line_code.isnot(None)
    ).group_by(Device.line_code).all()
    
    return {
        "total": total,
        "active": active,
        "inactive": inactive,
        "maintenance": maintenance,
        "enabled": enabled,
        "by_type": [{"type": t, "count": c} for t, c in type_stats],
        "by_line": [{"line": l, "count": c} for l, c in line_stats]
    }


@router.get("/download-template")
async def download_template():
    """下载设备导入模板（xlsx 格式）"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "设备导入模板"

    header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_alignment = Alignment(vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    headers = [
        '设备编号 *',
        '设备名称 *',
        '设备类型',
        '型号',
        '制造商',
        '产线编码',
        '工厂编码',
        '组编码',
        '描述',
        '安装位置',
        'IP地址',
        '状态'
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 30
    ws.column_dimensions['J'].width = 20
    ws.column_dimensions['K'].width = 18
    ws.column_dimensions['L'].width = 15

    ws.row_dimensions[1].height = 30

    example_data = [
        ['DEV-001', '注塑机A', '注塑机', 'JM-200T', '海天', 'L01', 'F01', 'G01', '1号注塑机', 'A区1楼', '192.168.1.100', 'active'],
        ['DEV-002', '组装线B', '组装线', 'AL-100', '自制', 'L02', 'F01', 'G02', '2号组装线', 'B区2楼', '192.168.1.101', 'active'],
    ]

    for row_num, row_data in enumerate(example_data, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border

    note_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    note_font = Font(name='微软雅黑', size=10, color='FF6600')

    ws.merge_cells('A5:L5')
    note_cell = ws.cell(row=5, column=1, value='说明：带 * 号为必填字段，状态只能填写 active、inactive 或 maintenance')
    note_cell.font = note_font
    note_cell.fill = note_fill

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = urllib.parse.quote('设备导入模板.xlsx')
    return Response(
        content=output.getvalue(),
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            'Content-Disposition': f"attachment; filename*=UTF-8''{filename}"
        }
    )


@router.post("/batch/toggle-dashboard")
async def batch_toggle_dashboard(
    device_codes: List[str],
    show: bool,
    db: Session = Depends(get_db)
):
    """批量切换设备是否在看板显示"""
    count = db.query(Device).filter(
        Device.device_code.in_(device_codes)
    ).update({
        Device.show_on_dashboard: show
    }, synchronize_session=False)

    db.commit()

    return {
        "message": f"成功{'显示' if show else '隐藏'} {count} 台设备",
        "show": show,
        "count": count
    }


@router.post("/import")
async def import_devices(
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """从 CSV 或 XLSX 文件导入设备数据"""
    if not (file.filename.endswith('.csv') or file.filename.endswith('.xlsx')):
        raise HTTPException(status_code=400, detail="仅支持 CSV 或 XLSX 格式文件")

    try:
        content = await file.read()
        rows_data = []

        if file.filename.endswith('.csv'):
            encodings_to_try = ['utf-8', 'gbk', 'gb2312', 'gb18030', 'utf-8-sig']
            csv_content = None
            for encoding in encodings_to_try:
                try:
                    csv_content = content.decode(encoding)
                    break
                except UnicodeDecodeError:
                    continue
            if csv_content is None:
                raise HTTPException(status_code=400, detail="无法识别的文件编码，请确保 CSV 文件使用 UTF-8 或 GBK 编码")
            reader = csv.DictReader(io.StringIO(csv_content))
            for row in reader:
                rows_data.append(row)
        else:
            wb = openpyxl.load_workbook(io.BytesIO(content))
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = {}
                for col_idx, value in enumerate(row):
                    if col_idx < len(headers):
                        row_dict[headers[col_idx]] = str(value) if value is not None else ''
                rows_data.append(row_dict)

        imported_count = 0
        updated_count = 0
        errors = []

        # 标准化表头：去掉 * 号和多余空格，兼容带 * 号的模板
        normalized_rows = []
        for row in rows_data:
            new_row = {}
            for key, value in row.items():
                clean_key = key.replace('*', '').strip() if key else key
                new_row[clean_key] = value
            normalized_rows.append(new_row)
        rows_data = normalized_rows

        for row_num, row in enumerate(rows_data, start=2):
            try:
                device_code = row.get('设备编号', '').strip()
                device_name = row.get('设备名称', '').strip()

                if not device_code or not device_name:
                    errors.append(f"第{row_num}行：缺少必填字段（设备编号或设备名称）")
                    continue

                status = row.get('状态', 'active').strip()
                if status not in ('active', 'inactive', 'maintenance', 'unknown'):
                    status = 'active'

                existing = db.query(Device).filter(Device.device_code == device_code).first()

                if existing:
                    existing.device_name = device_name
                    existing.device_type = row.get('设备类型', '').strip() or None
                    existing.model = row.get('型号', '').strip() or None
                    existing.manufacturer = row.get('制造商', '').strip() or None
                    existing.line_code = row.get('产线编码', '').strip() or None
                    existing.factory_code = row.get('工厂编码', '').strip() or None
                    existing.group_code = row.get('组编码', '').strip() or None
                    existing.description = row.get('描述', '').strip() or None
                    existing.location = row.get('安装位置', '').strip() or None
                    existing.ip_address = row.get('IP地址', '').strip() or None
                    existing.status = status
                    existing.updated_at = datetime.now()
                    updated_count += 1
                else:
                    device = Device(
                        device_code=device_code,
                        device_name=device_name,
                        device_type=row.get('设备类型', '').strip() or None,
                        model=row.get('型号', '').strip() or None,
                        manufacturer=row.get('制造商', '').strip() or None,
                        line_code=row.get('产线编码', '').strip() or None,
                        factory_code=row.get('工厂编码', '').strip() or None,
                        group_code=row.get('组编码', '').strip() or None,
                        description=row.get('描述', '').strip() or None,
                        location=row.get('安装位置', '').strip() or None,
                        ip_address=row.get('IP地址', '').strip() or None,
                        status=status,
                    )
                    db.add(device)
                    imported_count += 1

            except Exception as e:
                errors.append(f"第{row_num}行：{str(e)}")

        db.commit()

        return {
            "message": "导入完成",
            "imported": imported_count,
            "updated": updated_count,
            "errors": errors[:10]
        }

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"导入失败：{str(e)}")


@router.get("/export")
async def export_devices(
    device_code: Optional[str] = None,
    device_name: Optional[str] = None,
    status: Optional[str] = None,
    device_type: Optional[str] = None,
    line_code: Optional[str] = None,
    factory_code: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """导出设备数据为 CSV 文件"""
    db_query = db.query(Device)

    if device_code:
        db_query = db_query.filter(Device.device_code.contains(device_code))
    if device_name:
        db_query = db_query.filter(Device.device_name.contains(device_name))
    if status:
        db_query = db_query.filter(Device.status == status)
    if device_type:
        db_query = db_query.filter(Device.device_type == device_type)
    if line_code:
        db_query = db_query.filter(Device.line_code == line_code)
    if factory_code:
        db_query = db_query.filter(Device.factory_code == factory_code)

    devices = db_query.order_by(Device.created_at.desc()).all()

    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow([
        '设备编号', '设备名称', '设备类型', '型号', '制造商',
        '产线编码', '工厂编码', '组编码', '描述', '安装位置',
        '状态', '是否启用', 'IP地址', 'MQTT Topics', '创建时间', '更新时间'
    ])

    for device in devices:
        writer.writerow([
            device.device_code,
            device.device_name,
            device.device_type or '',
            device.model or '',
            device.manufacturer or '',
            device.line_code or '',
            device.factory_code or '',
            device.group_code or '',
            device.description or '',
            device.location or '',
            device.status,
            '是' if device.is_enabled else '否',
            device.ip_address or '',
            device.mqtt_topics or '',
            device.created_at.strftime('%Y-%m-%d %H:%M:%S') if device.created_at else '',
            device.updated_at.strftime('%Y-%m-%d %H:%M:%S') if device.updated_at else ''
        ])

    csv_content = output.getvalue()
    filename = f"devices_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

    return StreamingResponse(
        iter([csv_content]),
        media_type='text/csv; charset=utf-8',
        headers={
            'Content-Disposition': f"attachment; filename*=UTF-8''{filename}"
        }
    )


@router.get("/alarm-data")
async def get_alarm_data(
    device_code: Optional[str] = None,
    alarm_level: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    page: int = 1,
    page_size: int = 20,
    db: Session = Depends(get_db)
):
    """获取 ALARM 数据"""
    db_query = db.query(AlarmRecord)
    
    if device_code:
        db_query = db_query.filter(AlarmRecord.device_code.contains(device_code))
    
    if alarm_level:
        db_query = db_query.filter(AlarmRecord.alarm_level == alarm_level)
    
    if start_date:
        db_query = db_query.filter(AlarmRecord.alarm_time >= start_date)
    
    if end_date:
        db_query = db_query.filter(AlarmRecord.alarm_time <= end_date)
    
    total = db_query.count()
    
    offset = (page - 1) * page_size
    items = db_query.order_by(AlarmRecord.alarm_time.desc()).offset(offset).limit(page_size).all()
    
    return {
        "total": total,
        "items": [
            {
                "id": item.id,
                "alarm_code": item.alarm_code,
                "alarm_level": item.alarm_level,
                "alarm_type": item.alarm_type,
                "title": item.title,
                "description": item.description,
                "device_code": item.device_code,
                "device_name": item.device_name,
                "alarm_value": item.alarm_value,
                "threshold_value": item.threshold_value,
                "status": item.status,
                "handler": item.handler,
                "handled_at": item.handled_at.isoformat() if item.handled_at else None,
                "handle_remark": item.handle_remark,
                "alarm_time": item.alarm_time.isoformat() if item.alarm_time else None,
                "created_at": item.created_at.isoformat() if item.created_at else None
            }
            for item in items
        ]
    }


@router.get("/pv-data")
async def get_pv_data(
    device_code: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    page: int = 1,
    page_size: int = 20,
    db: Session = Depends(get_db)
):
    """获取 PV 数据"""
    db_query = db.query(PVCompressedParam)
    
    if device_code:
        db_query = db_query.filter(PVCompressedParam.topic.contains(device_code))
    
    if start_date:
        db_query = db_query.filter(PVCompressedParam.timestamp >= start_date)
    
    if end_date:
        db_query = db_query.filter(PVCompressedParam.timestamp <= end_date)
    
    total = db_query.count()
    
    offset = (page - 1) * page_size
    items = db_query.order_by(PVCompressedParam.timestamp.desc()).offset(offset).limit(page_size).all()
    
    return {
        "total": total,
        "items": [
            {
                "id": item.id,
                "topic": item.topic,
                "event_uid": item.event_uid,
                "timestamp": item.timestamp.isoformat() if item.timestamp else None,
                "original_timestamp": item.original_timestamp.isoformat() if item.original_timestamp else None,
                "created_at": item.created_at.isoformat() if item.created_at else None
            }
            for item in items
        ]
    }


@router.get("/sv-data")
async def get_sv_data(
    device_code: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    page: int = 1,
    page_size: int = 20,
    db: Session = Depends(get_db)
):
    """获取 SV 数据"""
    db_query = db.query(SVCompressedParam)
    
    if device_code:
        db_query = db_query.filter(SVCompressedParam.topic.contains(device_code))
    
    if start_date:
        db_query = db_query.filter(SVCompressedParam.timestamp >= start_date)
    
    if end_date:
        db_query = db_query.filter(SVCompressedParam.timestamp <= end_date)
    
    total = db_query.count()
    
    offset = (page - 1) * page_size
    items = db_query.order_by(SVCompressedParam.timestamp.desc()).offset(offset).limit(page_size).all()
    
    return {
        "total": total,
        "items": [
            {
                "id": item.id,
                "topic": item.topic,
                "event_uid": item.event_uid,
                "timestamp": item.timestamp.isoformat() if item.timestamp else None,
                "original_timestamp": item.original_timestamp.isoformat() if item.original_timestamp else None,
                "created_at": item.created_at.isoformat() if item.created_at else None
            }
            for item in items
        ]
    }


@router.get("/event-data")
async def get_event_data(
    device_code: Optional[str] = None,
    event_type: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    page: int = 1,
    page_size: int = 20,
    db: Session = Depends(get_db)
):
    """获取 Event 数据"""
    db_query = db.query(EventData)
    
    if device_code:
        db_query = db_query.filter(EventData.machine_id.contains(device_code))
    
    if start_date:
        start_ts = int(datetime.fromisoformat(start_date).timestamp() * 1000)
        db_query = db_query.filter(EventData.start_time >= start_ts)
    
    if end_date:
        end_ts = int(datetime.fromisoformat(end_date).timestamp() * 1000)
        db_query = db_query.filter(EventData.start_time <= end_ts)
    
    total = db_query.count()
    
    offset = (page - 1) * page_size
    items = db_query.order_by(EventData.start_time.desc()).offset(offset).limit(page_size).all()
    
    return {
        "total": total,
        "items": [
            {
                "id": item.id,
                "event_uid": item.event_uid,
                "start_code": item.start_code,
                "skin_code": item.skin_code,
                "start_time": item.start_time,
                "end_time": item.end_time,
                "duringtime": item.duringtime,
                "machine_duringtime": item.machine_duringtime,
                "machine_id": item.machine_id,
                "operator_id": item.operator_id,
                "operator_name": item.operator_name,
                "group_code": item.group_code,
                "group_name": item.group_name,
                "group_short_name": item.group_short_name,
                "factory_code": item.factory_code,
                "factory_name": item.factory_name,
                "line_code": item.line_code,
                "process_no": item.process_no,
                "created_at": item.created_at.isoformat() if item.created_at else None
            }
            for item in items
        ]
    }


@router.get("/compressed-data/latest")
async def get_latest_compressed_data(
    topic: str = Query(..., description="Topic 名称"),
    db: Session = Depends(get_db)
):
    """
    获取指定topic的最新一条压缩数据并解压
    
    Args:
        topic: MQTT Topic 名称
    
    Returns:
        最新一条解压后的数据，包含时间戳和完整数据内容
    """
    topic_config = db.query(MQTTTopicConfig).filter(
        MQTTTopicConfig.topic_name == topic
    ).first()
    
    if not topic_config:
        raise HTTPException(status_code=404, detail="Topic 配置不存在")
    
    latest_record = None
    
    if topic_config.topic_type in ('pv_compress', 'pv'):
        latest_record = db.query(PVCompressedParam).filter(
            PVCompressedParam.topic == topic
        ).order_by(PVCompressedParam.timestamp.desc()).first()
    elif topic_config.topic_type in ('sv_compress', 'sv'):
        latest_record = db.query(SVCompressedParam).filter(
            SVCompressedParam.topic == topic
        ).order_by(SVCompressedParam.timestamp.desc()).first()
    elif topic_config.topic_type in ('alarm_compress', 'alarm'):
        from app.models.alarm_compressed_param import AlarmCompressedParam
        latest_record = db.query(AlarmCompressedParam).filter(
            AlarmCompressedParam.topic == topic
        ).order_by(AlarmCompressedParam.timestamp.desc()).first()
    elif topic_config.topic_type == 'event':
        # 通过 topic_config.device_code 获取设备编号
        if topic_config.device_code:
            device = db.query(Device).filter(Device.device_code == topic_config.device_code).first()
            if device:
                latest_record = db.query(EventData).filter(
                    EventData.machine_id == device.device_code
                ).order_by(EventData.start_time.desc()).first()
        else:
            # 如果没有 device_code，尝试用 topic 名称匹配 machine_id（兼容旧数据）
            latest_record = db.query(EventData).filter(
                EventData.machine_id == topic
            ).order_by(EventData.start_time.desc()).first()
    else:
        # 自定义类型或其他类型，尝试从MQTT缓冲区获取最新数据
        from app.services.mqtt_collector import get_collector
        collector = get_collector()
        mqtt_buffer = collector.data_buffer if collector else []
        mqtt_msg = None
        for msg in reversed(mqtt_buffer):
            if msg.get('topic') == topic:
                mqtt_msg = msg
                break
        if mqtt_msg:
            return {
                "topic": topic,
                "topic_type": topic_config.topic_type,
                "has_data": True,
                "data": mqtt_msg.get('payload', {}),
                "timestamp": mqtt_msg.get('timestamp'),
                "source": "mqtt_buffer"
            }
        return {
            "topic": topic,
            "topic_type": topic_config.topic_type,
            "has_data": False,
            "data": None,
            "timestamp": None,
            "source": "none"
        }
    
    if not latest_record:
        return {
            "topic": topic,
            "topic_type": topic_config.topic_type,
            "has_data": False,
            "data": None
        }
    
    try:
        if topic_config.topic_type == 'event':
            event_data = {
                "event_uid": latest_record.event_uid,
                "start_code": latest_record.start_code,
                "skin_code": latest_record.skin_code,
                "start_time": latest_record.start_time,
                "end_time": latest_record.end_time,
                "start_signal": latest_record.start_signal,
                "end_signal": latest_record.end_signal,
                "duringtime": latest_record.duringtime,
                "machine_duringtime": latest_record.machine_duringtime,
                "machine_id": latest_record.machine_id,
                "operator_id": latest_record.operator_id,
                "operator_name": latest_record.operator_name,
                "group_code": latest_record.group_code,
                "group_name": latest_record.group_name,
                "group_short_name": latest_record.group_short_name,
                "factory_code": latest_record.factory_code,
                "factory_name": latest_record.factory_name,
                "line_code": latest_record.line_code,
                "process_no": latest_record.process_no,
            }
            if latest_record.extra_data:
                event_data["extra_data"] = json.loads(latest_record.extra_data)
            
            return {
                "topic": topic,
                "topic_type": topic_config.topic_type,
                "has_data": True,
                "timestamp": datetime.fromtimestamp(latest_record.start_time / 1000).isoformat() if latest_record.start_time else None,
                "original_timestamp": datetime.fromtimestamp(latest_record.start_time / 1000).isoformat() if latest_record.start_time else None,
                "data": event_data,
                "parse_rules": json.loads(topic_config.parse_rules) if topic_config.parse_rules else None
            }
        else:
            decompressed_data = decompress_payload(latest_record.compressed_payload)
            
            return {
                "topic": topic,
                "topic_type": topic_config.topic_type,
                "has_data": True,
                "timestamp": latest_record.timestamp.isoformat() if latest_record.timestamp else None,
                "original_timestamp": latest_record.original_timestamp.isoformat() if latest_record.original_timestamp else None,
                "data": decompressed_data,
                "parse_rules": json.loads(topic_config.parse_rules) if topic_config.parse_rules else None
            }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"解压缩失败：{str(e)}")


@router.get("/{device_code}", response_model=DeviceDetail)
async def get_device(device_code: str, db: Session = Depends(get_db)):
    """获取设备详情"""
    device = db.query(Device).filter(Device.device_code == device_code).first()
    if not device:
        raise HTTPException(status_code=404, detail="设备不存在")
    
    device_dict = parse_device_mqtt_topics(device)
    return DeviceDetail(**device_dict)


@router.post("/", response_model=DeviceDetail)
async def create_device(device: DeviceCreate, db: Session = Depends(get_db)):
    """创建新设备"""
    # 检查设备编号是否已存在
    existing = db.query(Device).filter(Device.device_code == device.device_code).first()
    if existing:
        raise HTTPException(status_code=400, detail="设备编号已存在")
    
    device_data = device.model_dump()
    # 将 mqtt_topics 列表转换为 JSON 字符串
    if 'mqtt_topics' in device_data and device_data['mqtt_topics'] is not None:
        device_data['mqtt_topics'] = json.dumps(device_data['mqtt_topics'])
    
    db_device = Device(**device_data)
    db.add(db_device)
    db.commit()
    db.refresh(db_device)
    
    # 同步更新 topic 配置的 device_code
    if device.mqtt_topics:
        for topic_name in device.mqtt_topics:
            topic_config = db.query(MQTTTopicConfig).filter(
                MQTTTopicConfig.topic_name == topic_name
            ).first()
            if topic_config and topic_config.device_code != db_device.device_code:
                topic_config.device_code = db_device.device_code
        
        db.commit()
    
    device_dict = parse_device_mqtt_topics(db_device)
    return DeviceDetail(**device_dict)


@router.put("/{device_code}", response_model=DeviceDetail)
async def update_device(device_code: str, device: DeviceUpdate, db: Session = Depends(get_db)):
    """更新设备信息"""
    db_device = db.query(Device).filter(Device.device_code == device_code).first()
    if not db_device:
        raise HTTPException(status_code=404, detail="设备不存在")
    
    # 记录旧的 mqtt_topics
    old_topics = []
    if db_device.mqtt_topics:
        try:
            old_topics = json.loads(db_device.mqtt_topics)
        except (json.JSONDecodeError, TypeError):
            old_topics = []
    
    # 更新字段
    update_data = device.model_dump(exclude_unset=True)
    
    # 移除 device_code 字段（编辑时不允许修改设备编号）
    update_data.pop('device_code', None)
    
    # 将 mqtt_topics 列表转换为 JSON 字符串
    if 'mqtt_topics' in update_data and update_data['mqtt_topics'] is not None:
        update_data['mqtt_topics'] = json.dumps(update_data['mqtt_topics'])
    
    for field, value in update_data.items():
        setattr(db_device, field, value)
    
    db_device.updated_at = datetime.now()
    db.commit()
    db.refresh(db_device)
    
    # 同步更新 topic 配置的 device_code
    new_topics = []
    if db_device.mqtt_topics:
        try:
            new_topics = json.loads(db_device.mqtt_topics)
        except (json.JSONDecodeError, TypeError):
            new_topics = []
    
    # 找出新增的 topic
    added_topics = [t for t in new_topics if t not in old_topics]
    # 找出移除的 topic
    removed_topics = [t for t in old_topics if t not in new_topics]
    
    # 更新新增 topic 的 device_code
    for topic_name in added_topics:
        topic_config = db.query(MQTTTopicConfig).filter(
            MQTTTopicConfig.topic_name == topic_name
        ).first()
        if topic_config and topic_config.device_code != device_code:
            topic_config.device_code = device_code
    
    # 更新移除 topic 的 device_code（设为 null）
    for topic_name in removed_topics:
        topic_config = db.query(MQTTTopicConfig).filter(
            MQTTTopicConfig.topic_name == topic_name
        ).first()
        if topic_config and topic_config.device_code == device_code:
            topic_config.device_code = None
    
    if added_topics or removed_topics:
        db.commit()
    
    device_dict = parse_device_mqtt_topics(db_device)
    return DeviceDetail(**device_dict)


@router.delete("/{device_code}")
async def delete_device(device_code: str, db: Session = Depends(get_db)):
    """删除设备"""
    device = db.query(Device).filter(Device.device_code == device_code).first()
    if not device:
        raise HTTPException(status_code=404, detail="设备不存在")
    
    # 同步清除 topic 配置的 device_code
    if device.mqtt_topics:
        try:
            topics = json.loads(device.mqtt_topics)
            for topic_name in topics:
                topic_config = db.query(MQTTTopicConfig).filter(
                    MQTTTopicConfig.topic_name == topic_name
                ).first()
                if topic_config and topic_config.device_code == device_code:
                    topic_config.device_code = None
        except (json.JSONDecodeError, TypeError):
            pass
    
    db.delete(device)
    db.commit()
    return {"message": "设备已删除"}


@router.get("/{device_code}/data-collection-config")
async def get_device_data_collection_config(
    device_code: str,
    db: Session = Depends(get_db)
):
    """
    获取设备数据采集配置

    根据设备编号获取其关联的topic配置，以及每个topic的预设键名列表
    用于前端动态生成数据采集展示页面

    Returns:
        设备的topic配置列表，包含topic名称、类型、键名定义等
    """
    device = db.query(Device).filter(Device.device_code == device_code).first()
    if not device:
        raise HTTPException(status_code=404, detail="设备不存在")
    
    device_dict = parse_device_mqtt_topics(device)
    
    topic_configs = []
    
    for topic_name in device_dict.get('mqtt_topics', []):
        topic_config = db.query(MQTTTopicConfig).filter(
            MQTTTopicConfig.topic_name == topic_name
        ).first()
        
        if topic_config:
            topic_configs.append({
                "topic_name": topic_config.topic_name,
                "topic_type": topic_config.topic_type,
                "description": topic_config.description,
                "enabled": topic_config.enabled,
                "parse_rules": json.loads(topic_config.parse_rules) if topic_config.parse_rules else None
            })
    
    return {
        "device_code": device.device_code,
        "device_name": device.device_name,
        "topic_configs": topic_configs
    }


@router.get("/{device_code}/data-source/status")
async def get_device_data_source_status(
    device_code: str,
    db: Session = Depends(get_db)
):
    """
    获取设备数据源状态

    返回MQTT连接状态、数据库连接状态以及Topic配置信息

    Args:
        device_code: 设备编号

    Returns:
        数据源状态信息，包括MQTT和数据库的连接状态
    """
    device = db.query(Device).filter(Device.device_code == device_code).first()
    if not device:
        raise HTTPException(status_code=404, detail="设备不存在")
    
    collector = get_device_data_collector()
    status = collector.get_data_source_status(device_code)
    
    return status


@router.get("/{device_code}/data-source/latest")
async def get_device_latest_data(
    device_code: str,
    topic_name: str = Query(..., description="Topic名称"),
    source: str = Query("auto", description="数据源：auto/mqtt/database"),
    db: Session = Depends(get_db)
):
    """
    获取设备最新数据

    支持从MQTT实时获取或从数据库读取

    Args:
        device_code: 设备编号
        topic_name: Topic名称
        source: 数据源类型
            - auto: 自动选择（优先MQTT，失败则数据库）
            - mqtt: 仅从MQTT获取
            - database: 仅从数据库获取
            
    Returns:
        最新数据，包含数据源标识
    """
    device = db.query(Device).filter(Device.device_code == device_code).first()
    if not device:
        raise HTTPException(status_code=404, detail="设备不存在")

    collector = get_device_data_collector()

    if source == "mqtt":
        data = collector.get_realtime_data_from_mqtt(device_code, topic_name)
        if not data:
            raise HTTPException(status_code=404, detail="MQTT实时数据不可用")
        return data

    elif source == "database":
        data = collector.get_latest_data_from_db(device_code, topic_name, db)
        if not data:
            raise HTTPException(status_code=404, detail="数据库无数据")
        return data

    else:  # auto
        # 优先尝试MQTT
        data = collector.get_realtime_data_from_mqtt(device_code, topic_name)
        if data:
            return data

        # 降级到数据库
        data = collector.get_latest_data_from_db(device_code, topic_name, db)
        if data:
            return data

        raise HTTPException(status_code=404, detail="无可用数据")


@router.get("/{device_code}/data-source/historical")
async def get_device_historical_data(
    device_code: str,
    topic_name: str = Query(..., description="Topic名称"),
    start_time: Optional[str] = Query(None, description="开始时间"),
    end_time: Optional[str] = Query(None, description="结束时间"),
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(20, ge=1, le=100, description="每页数量"),
    db: Session = Depends(get_db)
):
    """
    获取设备历史数据（从数据库）

    Args:
        device_code: 设备编号
        topic_name: Topic名称
        start_time: 开始时间（ISO格式）
        end_time: 结束时间（ISO格式）
        page: 页码
        page_size: 每页数量

    Returns:
        历史数据列表（分页）
    """
    device = db.query(Device).filter(Device.device_code == device_code).first()
    if not device:
        raise HTTPException(status_code=404, detail="设备不存在")

    collector = get_device_data_collector()
    data = collector.get_historical_data(
        device_code=device_code,
        topic_name=topic_name,
        start_time=start_time,
        end_time=end_time,
        page=page,
        page_size=page_size
    )
    
    return data


@router.post("/{device_code}/data-source/subscribe")
async def subscribe_device_realtime_data(
    device_code: str,
    topic_name: str = Query(..., description="Topic名称"),
    db: Session = Depends(get_db)
):
    """
    订阅设备实时数据（通过MQTT）

    Args:
        device_code: 设备编号
        topic_name: Topic名称
        
    Returns:
        订阅结果
    """
    device = db.query(Device).filter(Device.device_code == device_code).first()
    if not device:
        raise HTTPException(status_code=404, detail="设备不存在")

    collector = get_device_data_collector()

    # 这里可以添加WebSocket推送逻辑
    # 目前仅返回订阅状态
    success = collector.subscribe_realtime(
        device_code=device_code,
        topic_name=topic_name,
        callback=lambda data: logger.info(f"收到实时数据：{data}")
    )

    if success:
        return {
            "message": "订阅成功",
            "device_code": device_code,
            "topic_name": topic_name,
            "status": "subscribed"
        }
    else:
        raise HTTPException(status_code=500, detail="订阅失败")


@router.get("/{device_code}/mqtt/latest")
async def get_latest_mqtt_message(
    device_code: str,
    topic_name: str = Query(..., description="Topic名称"),
    db: Session = Depends(get_db)
):
    """
    获取设备指定Topic的最新MQTT消息

    Args:
        device_code: 设备编号
        topic_name: Topic名称

    Returns:
        最新的MQTT消息数据
    """
    device = db.query(Device).filter(Device.device_code == device_code).first()
    if not device:
        raise HTTPException(status_code=404, detail="设备不存在")
    
    mqtt_collector = get_collector()
    if not mqtt_collector:
        raise HTTPException(status_code=500, detail="MQTT Collector未初始化")
        
    mqtt_buffer = mqtt_collector.data_buffer if mqtt_collector else []
    
    # 查找匹配topic的最新消息
    latest_msg = None
    for message in reversed(mqtt_buffer):
        if message.get('topic') == topic_name:
            latest_msg = message
            break
            
    if not latest_msg:
        raise HTTPException(status_code=404, detail="未找到该Topic的最新消息")
        
    return {
        "device_code": device_code,
        "topic_name": topic_name,
        "message": latest_msg
    }


@router.get("/{device_code}/monitor/status")
async def get_device_monitor_status(
    device_code: str,
    db: Session = Depends(get_db)
):
    """
    获取设备监控状态

    从MQTT实时消息中根据配置规则获取设备状态信息，包括：
    - 加工状态
    - 换模状态
    - 故障状态
    - 报警状态
    - 缺料状态
    - 停机状态
    - 当前加工产品零件号
    - U9订单信息
    """
    device = db.query(Device).filter(Device.device_code == device_code).first()
    if not device:
        raise HTTPException(status_code=404, detail="设备不存在")
    
    mqtt_topics = []
    if device.mqtt_topics:
        try:
            mqtt_topics = json.loads(device.mqtt_topics)
        except Exception as e:
            mqtt_topics = []
    
    topic_configs = db.query(MQTTTopicConfig).filter(
        MQTTTopicConfig.topic_name.in_(mqtt_topics)
    ).all()
    
    now = datetime.now()
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    today_start_ts = int(today_start.timestamp() * 1000)
    
    mqtt_collector = get_collector()
    mqtt_buffer = mqtt_collector.data_buffer if mqtt_collector else []

    def get_latest_mqtt_message(topic_name: str) -> Optional[Dict[str, Any]]:
        # 先从MQTT缓冲区查找
        for message in reversed(mqtt_buffer):
            if message.get('topic') == topic_name:
                return message

        # 缓冲区没有，从数据库查找（pv/sv/alarm压缩数据）
        topic_config = next((tc for tc in topic_configs if tc.topic_name == topic_name), None)
        if topic_config and topic_config.topic_type in ('pv_compress', 'pv', 'sv_compress', 'sv', 'alarm_compress', 'alarm'):
            try:
                if topic_config.topic_type in ('pv_compress', 'pv'):
                    from app.models.pv_compressed_param import PVCompressedParam
                    record = db.query(PVCompressedParam).filter(
                        PVCompressedParam.topic == topic_name
                    ).order_by(PVCompressedParam.timestamp.desc()).first()
                elif topic_config.topic_type in ('sv_compress', 'sv'):
                    from app.models.sv_compressed_param import SVCompressedParam
                    record = db.query(SVCompressedParam).filter(
                        SVCompressedParam.topic == topic_name
                    ).order_by(SVCompressedParam.timestamp.desc()).first()
                else:
                    from app.models.alarm_compressed_param import AlarmCompressedParam
                    record = db.query(AlarmCompressedParam).filter(
                        AlarmCompressedParam.topic == topic_name
                    ).order_by(AlarmCompressedParam.timestamp.desc()).first()

                if record and record.compressed_payload:
                    decompressed = gzip.decompress(record.compressed_payload).decode('utf-8')
                    payload = json.loads(decompressed)
                    return {
                        'topic': topic_name,
                        'payload': payload,
                        'timestamp': str(record.timestamp)
                    }
            except Exception as e:
                print(f"[规则评估] 从数据库获取压缩数据失败 topic={topic_name}: {e}")

        return None

    def evaluate_single_condition(condition: Dict[str, Any]) -> tuple:
        """评估单个条件是否匹配，返回 (bool, detail_dict)"""
        topic_name = condition.get('topic_name')
        field_path = condition.get('field_path')
        match_rule = condition.get('match_rule')
        match_value_str = condition.get('match_value')
        extraction_rule = condition.get('extraction_rule')

        detail = {
            "topic_name": topic_name,
            "field_path": field_path,
            "match_rule": match_rule,
            "match_value": match_value_str,
            "steps": []
        }

        msg = get_latest_mqtt_message(topic_name)
        if not msg:
            detail["steps"].append({"step": "获取消息", "result": "失败", "reason": f"未找到 topic={topic_name} 的数据"})
            detail["final_result"] = False
            print(f"  [条件诊断] topic={topic_name}: 未找到消息")
            return False, detail

        payload = msg.get('payload', {})
        actual_value = _get_nested_value(payload, field_path)
        detail["steps"].append({"step": "获取字段值", "field_path": field_path, "actual_value": str(actual_value)[:200], "type": type(actual_value).__name__})
        print(f"  [条件诊断] topic={topic_name}, field={field_path}, value={repr(actual_value)[:100]}, type={type(actual_value).__name__}")

        if actual_value is None:
            # 展示 payload 结构帮助诊断
            payload_keys = list(payload.keys()) if isinstance(payload, dict) else []
            detail["steps"].append({"step": "字段值检查", "result": "失败", "reason": f"字段 '{field_path}' 值为 None", "payload_keys": payload_keys[:30]})
            print(f"  [条件诊断] 字段 '{field_path}' 值为 None, payload keys: {payload_keys[:20]}")
            detail["final_result"] = False
            print(f"  [条件诊断] 字段 '{field_path}' 值为 None")
            return False, detail

        if extraction_rule and extraction_rule.get('type'):
            before = actual_value
            actual_value = _apply_extraction(actual_value, extraction_rule)
            detail["steps"].append({"step": "截取规则", "before": str(before)[:100], "after": str(actual_value)[:100]})
            print(f"  [条件诊断] 截取: {repr(before)[:60]} -> {repr(actual_value)[:60]}")
            if actual_value is None:
                detail["steps"].append({"step": "截取后检查", "result": "失败", "reason": "截取后值为 None"})
                detail["final_result"] = False
                return False, detail

        result = _match_value(actual_value, match_rule, match_value_str)
        detail["steps"].append({"step": "值匹配", "actual": str(actual_value)[:100], "rule": match_rule, "expected": str(match_value_str)[:100], "result": "匹配" if result else "不匹配"})
        detail["final_result"] = result
        print(f"  [条件诊断] 匹配: {repr(actual_value)[:60]} {match_rule} {repr(match_value_str)[:60]} => {'匹配' if result else '不匹配'}")
        return result, detail
    
    def evaluate_config(config) -> tuple:
        """评估一个配置规则是否匹配，支持多条件组合。返回 (bool, details)"""
        conditions = config.conditions
        logic_operator = config.logic_operator or 'AND'

        if not conditions:
            # 单条件路径：使用 config 自身字段
            msg = get_latest_mqtt_message(config.topic_name)
            if not msg:
                print(f"[规则评估] 配置#{config.id} ({config.status_type}): 未找到MQTT消息 topic={config.topic_name}")
                detail = {"topic_name": config.topic_name, "field_path": config.field_path,
                          "match_rule": config.match_rule, "match_value": config.match_value,
                          "steps": [{"step": "获取消息", "result": "失败", "reason": f"未找到 topic={config.topic_name} 的数据"}],
                          "final_result": False}
                return False, [detail]
            payload = msg.get('payload', {})
            actual_value = _get_nested_value(payload, config.field_path)
            print(f"[规则评估] 配置#{config.id} ({config.status_type}):")
            print(f"  - Topic: {config.topic_name}")
            print(f"  - 字段路径: {config.field_path}")
            print(f"  - 实际值: {actual_value} (类型: {type(actual_value).__name__})")
            print(f"  - 匹配规则: {config.match_rule}")
            print(f"  - 匹配值: {config.match_value} (原始)")

            detail = {"topic_name": config.topic_name, "field_path": config.field_path,
                      "match_rule": config.match_rule, "match_value": config.match_value, "steps": []}

            if actual_value is None:
                print(f"  - 结果: ❌ 字段值为None")
                detail["steps"].append({"step": "字段值检查", "result": "失败", "reason": "字段值为 None"})
                detail["final_result"] = False
                return False, [detail]
            if config.extraction_rule and config.extraction_rule.get('type'):
                actual_value = _apply_extraction(actual_value, config.extraction_rule)
                print(f"  - 截取后值: {actual_value} (类型: {type(actual_value).__name__})")
                if actual_value is None:
                    print(f"  - 结果: ❌ 截取后值为None")
                    detail["steps"].append({"step": "截取后检查", "result": "失败", "reason": "截取后值为 None"})
                    detail["final_result"] = False
                    return False, [detail]

            result = match_value(actual_value, config.match_rule, config.match_value)
            print(f"  - 匹配结果: {'✅ 成功' if result else '❌ 失败'}")
            detail["steps"].append({"step": "值匹配", "actual": str(actual_value)[:100], "rule": config.match_rule, "expected": str(config.match_value)[:100], "result": "匹配" if result else "不匹配"})
            detail["final_result"] = result
            return result, [detail]

        # 多条件路径
        condition_results = []
        condition_details = []
        for condition in conditions:
            result, detail = evaluate_single_condition(condition)
            condition_results.append(result)
            condition_details.append(detail)

        if logic_operator.upper() == 'OR':
            final = any(condition_results)
        else:
            final = all(condition_results)

        return final, condition_details
    
    status_configs = db.query(DeviceStatusMonitorConfig).filter(
        and_(
            DeviceStatusMonitorConfig.device_code == device_code,
            DeviceStatusMonitorConfig.enabled == True
        )
    ).order_by(DeviceStatusMonitorConfig.priority).all()

    # 加载基础规则 + 设备规则（设备规则覆盖同类型基础规则）
    config_by_status_type = load_device_rules(db, device_code, DeviceStatusMonitorConfig)
    
    status_result = {
        "processing": False,
        "stop": False,
        "fault_stop": False,
        "emergency stop": False,
        "mold_change": False,
        "maintain": False,
        "alarm": False,
        "material_shortage": False,
        "plan_stop": False
    }
    
    # 当前活跃状态列表，直观显示所有匹配到的状态
    active_statuses = []
    
    # 规则匹配详情，用于前端展示和调试
    rule_match_details = {}
    status_errors = []
    
    # 先获取事件消息和零件号（用于后续的伺服曲线匹配）
    latest_event_message = None
    event_topics = [tc for tc in topic_configs if tc.topic_type == 'event']
    for event_topic in event_topics:
        msg = get_latest_mqtt_message(event_topic.topic_name)
        if msg:
            if not latest_event_message or msg.get('timestamp', '') > latest_event_message.get('timestamp', ''):
                latest_event_message = msg
    
    current_part_number = None
    current_u9_material_code = None
    current_start_code = None
    
    # 首先尝试从配置中获取零件号
    from app.models.current_product_config import CurrentProductConfig
    product_configs = db.query(CurrentProductConfig).filter(
        and_(
            CurrentProductConfig.device_code == device_code,
            CurrentProductConfig.enabled == True
        )
    ).order_by(CurrentProductConfig.priority).all()
    
    if product_configs:
        # 使用配置的规则获取零件号
        for config in product_configs:
            try:
                msg = get_latest_mqtt_message(config.topic_name)
                if msg:
                    payload = msg.get('payload', {})
                    field_value = _get_nested_value(payload, config.field_path)
                    
                    if field_value:
                        current_start_code = str(field_value)
                        
                        extracted_value = _apply_extraction(current_start_code, config.extraction_rule) if config.extraction_rule else current_start_code
                        
                        # 尝试匹配物料
                        product = db.query(Material).filter(
                            or_(
                                Material.part_number == extracted_value,
                                Material.u9_material_code == extracted_value,
                                Material.specification == extracted_value
                            ),
                            Material.status == 'active'
                        ).first()
                        
                        if product:
                            current_part_number = product.part_number
                            current_u9_material_code = product.u9_material_code
                        else:
                            current_part_number = extracted_value
                            current_u9_material_code = extracted_value
                        
                        break  # 找到后跳出
            except Exception as e:
                print(f"从配置获取零件号时出错 (config_id={config.id}): {e}")
    
    # 如果配置中没有获取到，尝试从event消息中获取
    if not current_part_number and latest_event_message:
        try:
            latest_event_payload = latest_event_message.get('payload', {})
            event_start_code = latest_event_payload.get('start_code')
            
            if event_start_code:
                current_start_code = event_start_code
                
                # 尝试多种方式匹配物料
                product = db.query(Material).filter(
                    or_(
                        Material.part_number == event_start_code,
                        Material.u9_material_code == event_start_code,
                        Material.specification == event_start_code
                    ),
                    Material.status == 'active'
                ).first()
                
                if product:
                    current_part_number = product.part_number
                    current_u9_material_code = product.u9_material_code
                else:
                    current_part_number = event_start_code
                    current_u9_material_code = event_start_code
        except Exception as e:
            print(f"从event消息获取零件号时出错: {e}")
    
    # 检查状态配置规则
    # 互斥状态：计划加工、计划停机、故障停机、紧急停机、换模、维护、缺料 只能存在一种
    # 可共存状态：报警 可以与以上任意状态同时出现
    mutually_exclusive_statuses = ["processing", "stop", "fault_stop", "emergency stop", "mold_change", "maintain", "material_shortage"]
    coexistent_statuses = ["alarm"]
    
    # 临时存储所有匹配到的状态
    matched_statuses = {}
    
    for status_type, configs in config_by_status_type.items():
        rule_match_details[status_type] = {
            "matched": False,
            "matched_config_id": None,
            "matched_config_desc": None,
            "evaluated_configs": [],
            "error": None
        }
        
        matched_config = None
        
        for config in configs:
            config_detail = {
                "config_id": config.id,
                "description": config.description or f"配置#{config.id}",
                "priority": config.priority,
                "conditions_count": len(config.conditions) if config.conditions else 1,
                "matched": False,
                "error": None
            }
            
            try:
                eval_result, eval_details = evaluate_config(config)
                config_detail["evaluation_details"] = eval_details
                if eval_result:
                    config_detail["matched"] = True
                    rule_match_details[status_type]["matched"] = True
                    rule_match_details[status_type]["matched_config_id"] = config.id
                    rule_match_details[status_type]["matched_config_desc"] = config.description

                    # 记录匹配到的配置
                    matched_config = config
                    break
            except Exception as e:
                error_msg = f"评估配置失败 (config_id={config.id}): {str(e)}"
                print(error_msg)
                config_detail["error"] = str(e)
                status_errors.append(error_msg)
            
            rule_match_details[status_type]["evaluated_configs"].append(config_detail)
        
        # 如果该状态类型有匹配的配置，记录到临时存储
        if matched_config:
            matched_statuses[status_type] = matched_config
    
    # 处理配置了参数曲线的规则（包含基础规则 + 设备规则）
    all_configs = []
    for configs in config_by_status_type.values():
        all_configs.extend(configs)
    configs_with_curves = [config for config in all_configs if config.curve_id is not None]
    if configs_with_curves:
        db_param_curves = db.query(DBParamCurve).filter(
            and_(
                DBParamCurve.device_code == device_code,
                DBParamCurve.enabled == 1,
                DBParamCurve.id.in_([config.curve_id for config in configs_with_curves])
            )
        ).all()

        curve_map = {curve.id: curve for curve in db_param_curves}

        # 获取曲线起始时间（从event消息）
        curve_start_time = None
        if latest_event_message:
            msg_timestamp = latest_event_message.get('timestamp', '')
            if msg_timestamp:
                try:
                    curve_start_time = datetime.fromisoformat(msg_timestamp)
                except (ValueError, TypeError) as e:
                    print(f"[曲线匹配] 解析event时间戳失败: {e}")

        for config in configs_with_curves:
            curve = curve_map.get(config.curve_id)
            if not curve:
                continue

            try:
                if curve_start_time:
                    result = evaluate_curve_config(
                        config=config,
                        curve=curve,
                        mqtt_buffer=mqtt_buffer,
                        curve_start_time=curve_start_time,
                        field_path=config.field_path
                    )
                else:
                    # 无event消息时，使用最新MQTT数据直接匹配
                    result = evaluate_curve_config_from_buffer(
                        config=config,
                        curve=curve,
                        mqtt_buffer=mqtt_buffer,
                        field_path=config.field_path
                    )

                if result['matched']:
                    status_type = result['status_type']
                    if status_type not in matched_statuses or config.priority < matched_statuses[status_type].priority:
                        matched_statuses[status_type] = config
                        print(f"[曲线匹配] 状态 {status_type} 匹配成功 score={result['score']}%")
            except Exception as e:
                print(f"[曲线匹配] 评估失败 config_id={config.id}: {e}")
    
    # 处理互斥状态：加工、停机、换模、故障 只能存在一种
    # 按照优先级选择最终状态（优先级数字越小优先级越高）
    matched_mutually_exclusive = []
    for status_type in mutually_exclusive_statuses:
        if status_type in matched_statuses:
            config = matched_statuses[status_type]
            matched_mutually_exclusive.append({
                "status_type": status_type,
                "config": config,
                "priority": config.priority
            })
    
    # 按优先级排序，选择优先级最高的
    if matched_mutually_exclusive:
        matched_mutually_exclusive.sort(key=lambda x: x["priority"])
        final_mutually_exclusive = matched_mutually_exclusive[0]
        status_result[final_mutually_exclusive["status_type"]] = True
    
    # 处理可共存状态：缺料、报警
    for status_type in coexistent_statuses:
        if status_type in matched_statuses:
            status_result[status_type] = True
    
    # 构建活跃状态列表
    status_type_labels = {
        "processing": "计划加工",
        "stop": "计划停机",
        "fault_stop": "故障停机",
        "emergency stop": "紧急停机",
        "mold_change": "换模",
        "maintain": "维护",
        "alarm": "报警",
        "material_shortage": "缺料",
        "plan_stop": "计划停机"
    }
    
    # 添加互斥状态中优先级最高的
    if matched_mutually_exclusive:
        final = matched_mutually_exclusive[0]
        config = final["config"]
        
        # 判断匹配方法：如果有curve_id则是曲线匹配，否则是规则匹配
        match_method = "curve" if (config and config.curve_id) else "rule"
        
        # 获取实际匹配值
        matched_value = None
        actual_value = None
        if config and not config.curve_id:
            try:
                msg = get_latest_mqtt_message(config.topic_name)
                if msg:
                    payload = msg.get('payload', {})
                    actual_value = _get_nested_value(payload, config.field_path)
                    matched_value = config.match_value
            except Exception as e:
                print(f"[规则评估] 获取MQTT消息失败: {e}")

        status_info = {
            "status_type": final["status_type"],
            "status_label": status_type_labels.get(final["status_type"], final["status_type"]),
            "config_id": config.id if config else None,
            "config_description": config.description or f"配置#{config.id}" if config else "系统自动判断",
            "priority": config.priority if config else 999,
            "matched_at": datetime.now().isoformat(),
            "match_method": match_method,
            "is_mutually_exclusive": True,
            "topic_name": config.topic_name if config else None,
            "field_path": config.field_path if config else None,
            "match_rule": config.match_rule if config else None,
            "matched_value": matched_value,
            "actual_value": str(actual_value) if actual_value is not None else None
        }
        
        # 如果是曲线匹配，添加匹配分数
        if config and config.curve_id:
            status_info["curve_id"] = config.curve_id
        
        active_statuses.append(status_info)
    
    # 添加可共存状态
    for status_type in coexistent_statuses:
        if status_type in matched_statuses:
            config = matched_statuses[status_type]
            match_method = "curve" if config.curve_id else "rule"
            
            # 获取实际匹配值
            matched_value = None
            actual_value = None
            if not config.curve_id:
                try:
                    msg = get_latest_mqtt_message(config.topic_name)
                    if msg:
                        payload = msg.get('payload', {})
                        actual_value = _get_nested_value(payload, config.field_path)
                        matched_value = config.match_value
                except Exception as e:
                    print(f"[规则评估] 获取MQTT消息失败: {e}")

            status_info = {
                "status_type": status_type,
                "status_label": status_type_labels.get(status_type, status_type),
                "config_id": config.id,
                "config_description": config.description or f"配置#{config.id}",
                "priority": config.priority,
                "matched_at": datetime.now().isoformat(),
                "match_method": match_method,
                "is_mutually_exclusive": False,
                "topic_name": config.topic_name,
                "field_path": config.field_path,
                "match_rule": config.match_rule,
                "matched_value": matched_value,
                "actual_value": str(actual_value) if actual_value is not None else None
            }
            
            if config.curve_id:
                status_info["curve_id"] = config.curve_id
            
            active_statuses.append(status_info)
    
    # 处理订单信息
    current_order_info = None
    latest_event_payload = latest_event_message.get('payload', {}) if latest_event_message else None
    
    if current_start_code and current_u9_material_code:
        today_str = today_start.strftime("%Y-%m-%d")
        orders = db.query(ProductOrder).filter(
            and_(
                ProductOrder.u9_material_code == current_u9_material_code,
                ProductOrder.query_date == today_str
            )
        ).all()

        if orders:
            total_planned = sum(o.planned_output or 0 for o in orders)
            current_order_info = {
                "planned_output": total_planned,
                "order_count": len(orders),
                "details": [
                    {
                        "doc_no": o.doc_no,
                        "item_code": o.item_code,
                        "product_qty": o.product_qty,
                        "total_complete_qty": o.total_complete_qty,
                        "doc_state": o.doc_state,
                        "mold_no": o.mold_no,
                        "line_code": o.line_code,
                        "start_date": o.start_date.isoformat() if o.start_date else None
                    }
                    for o in orders[:5]
                ]
            }
    
    alarm_count = 0
    alarm_topics = [tc for tc in topic_configs if tc.topic_type in ('alarm_compress', 'alarm')]
    for alarm_topic in alarm_topics:
        try:
            alarm_msg = get_latest_mqtt_message(alarm_topic.topic_name)
            if alarm_msg:
                alarm_payload = alarm_msg.get('payload', {})
                if isinstance(alarm_payload, list):
                    alarm_count = len(alarm_payload)
                elif isinstance(alarm_payload, dict):
                    alarms = alarm_payload.get('alarms', [])
                    if isinstance(alarms, list):
                        alarm_count = len(alarms)
        except Exception as e:
            print(f"[规则评估] 获取报警数据失败: {e}")

    # 处理停机状态：如果规则没有匹配到停机，则根据设备状态判断
    if not status_result.get('stop', False):
        stop_status = device.status == 'inactive'
        if latest_event_message:
            msg_timestamp = latest_event_message.get('timestamp', '')
            if msg_timestamp:
                try:
                    msg_time = datetime.fromisoformat(msg_timestamp)
                    if (now - msg_time).total_seconds() > 3600:
                        stop_status = True
                except Exception as e:
                    print(f"[规则评估] 解析事件时间戳失败: {e}")

        # 如果判断为停机，需要重新处理互斥状态
        if stop_status:
            # 将停机状态加入互斥状态列表
            matched_mutually_exclusive.append({
                "status_type": "stop",
                "config": None,
                "priority": 999
            })
            
            # 重新按优先级排序
            matched_mutually_exclusive.sort(key=lambda x: x["priority"])
            
            # 清除之前的互斥状态结果
            for status_type in mutually_exclusive_statuses:
                status_result[status_type] = False
            
            # 设置新的互斥状态
            final = matched_mutually_exclusive[0]
            status_result[final["status_type"]] = True
            
            # 重新构建活跃状态列表中的互斥状态部分
            active_statuses = [s for s in active_statuses if not s.get("is_mutually_exclusive", False)]
            
            if final["config"]:
                active_statuses.insert(0, {
                    "status_type": final["status_type"],
                    "status_label": status_type_labels.get(final["status_type"], final["status_type"]),
                    "config_id": final["config"].id,
                    "config_description": final["config"].description or f"配置#{final['config'].id}",
                    "priority": final["config"].priority,
                    "matched_at": datetime.now().isoformat(),
                    "match_method": "rule",
                    "is_mutually_exclusive": True
                })
            else:
                active_statuses.insert(0, {
                    "status_type": "stop",
                    "status_label": "停机",
                    "config_id": None,
                    "config_description": "设备状态为inactive或长时间无事件消息",
                    "priority": 999,
                    "matched_at": datetime.now().isoformat(),
                    "match_method": "device_status",
                    "is_mutually_exclusive": True
                })
    
    # 处理计划停机状态（可与其他状态共存）
    plan_stop_status = False
    if current_order_info and current_order_info.get('details'):
        for detail in current_order_info['details']:
            if detail.get('doc_state') in ['completed', 'closed']:
                plan_stop_status = True
                break
    
    status_result['plan_stop'] = plan_stop_status
    
    # 如果计划停机状态为True，添加到活跃状态列表
    if plan_stop_status:
        active_statuses.append({
            "status_type": "plan_stop",
            "status_label": status_type_labels.get("plan_stop", "计划停机"),
            "config_id": None,
            "config_description": "订单已完成或关闭",
            "priority": 999,
            "matched_at": datetime.now().isoformat(),
            "match_method": "order_status",
            "is_mutually_exclusive": False
        })
    
    # 检查是否有未匹配到规则的状态类型
    for status_type in status_result.keys():
        if status_type not in rule_match_details:
            rule_match_details[status_type] = {
                "matched": False,
                "matched_config_id": None,
                "matched_config_desc": None,
                "evaluated_configs": [],
                "error": "未配置该状态的监控规则"
            }

    # 自动同步匹配状态到设备表
    try:
        new_status_parts = []
        if active_statuses:
            for s in active_statuses:
                if s.get('is_mutually_exclusive'):
                    cfg = matched_statuses.get(s.get('status_type'))
                    if cfg and cfg.device_status:
                        new_status_parts.append(cfg.device_status)
                    else:
                        new_status_parts.append(s.get('status_type', ''))
            for s in active_statuses:
                if not s.get('is_mutually_exclusive'):
                    new_status_parts.append(s.get('status_type', ''))
        new_status = ','.join(new_status_parts) if new_status_parts else device.status
        if new_status != device.status:
            device.status = new_status
            device.updated_at = datetime.now()
            db.commit()
            print(f"[状态同步] 设备 {device.device_code} 状态更新: {device.status} -> {new_status}")
    except Exception as e:
        print(f"[状态同步] 写入设备状态失败: {e}")

    return {
        "device_code": device.device_code,
        "device_name": device.device_name,
        "status": status_result,
        "active_statuses": active_statuses,
        "current_part": {
            "part_number": current_part_number,
            "u9_material_code": current_u9_material_code,
            "start_code": current_start_code,
            "order_info": current_order_info
        },
        "alarm_count": alarm_count,
        "latest_event_time": latest_event_message.get('timestamp') if latest_event_message else None,
        "query_time": now.isoformat(),
        "data_source": "mqtt_realtime",
        "config_based": len(status_configs) > 0,
        "rule_match_details": rule_match_details,
        "status_errors": status_errors,
        "has_errors": len(status_errors) > 0
    }


@router.post("/{device_code}/sync-status")
async def sync_device_status(
    device_code: str,
    db: Session = Depends(get_db)
):
    """
    根据规则匹配结果同步设备状态

    评估设备的所有 DeviceStatusMonitorConfig 规则，根据匹配结果更新 Device.status：
    - processing → scheduled processing
    - stop → scheduled outage
    - fault_stop → maintenance
    - emergency stop → emergency stop
    - mold_change → mold_change
    - maintain → maintain
    - alarm（仅）→ 不变
    - material_shortage → material_shortage
    - 无匹配 → unknown
    """
    device = db.query(Device).filter(Device.device_code == device_code).first()
    if not device:
        raise HTTPException(status_code=404, detail="设备不存在")

    old_status = device.status

    # 获取 MQTT 缓冲区
    mqtt_collector = get_collector()
    mqtt_buffer = mqtt_collector.data_buffer if mqtt_collector else []

    # 获取设备关联的topic配置
    mqtt_topics = []
    if device.mqtt_topics:
        try:
            mqtt_topics = json.loads(device.mqtt_topics)
        except Exception as e:
            mqtt_topics = []
    topic_configs_list = db.query(MQTTTopicConfig).filter(
        MQTTTopicConfig.topic_name.in_(mqtt_topics)
    ).all() if mqtt_topics else []

    def get_latest_mqtt_message(topic_name: str) -> Optional[Dict[str, Any]]:
        # 先从MQTT缓冲区查找
        for message in reversed(mqtt_buffer):
            if message.get('topic') == topic_name:
                return message

        # 缓冲区没有，从数据库查找（pv/sv/alarm压缩数据）
        topic_config = next((tc for tc in topic_configs_list if tc.topic_name == topic_name), None)
        if topic_config and topic_config.topic_type in ('pv_compress', 'pv', 'sv_compress', 'sv', 'alarm_compress', 'alarm'):
            try:
                if topic_config.topic_type in ('pv_compress', 'pv'):
                    from app.models.pv_compressed_param import PVCompressedParam
                    record = db.query(PVCompressedParam).filter(
                        PVCompressedParam.topic == topic_name
                    ).order_by(PVCompressedParam.timestamp.desc()).first()
                elif topic_config.topic_type in ('sv_compress', 'sv'):
                    from app.models.sv_compressed_param import SVCompressedParam
                    record = db.query(SVCompressedParam).filter(
                        SVCompressedParam.topic == topic_name
                    ).order_by(SVCompressedParam.timestamp.desc()).first()
                else:
                    from app.models.alarm_compressed_param import AlarmCompressedParam
                    record = db.query(AlarmCompressedParam).filter(
                        AlarmCompressedParam.topic == topic_name
                    ).order_by(AlarmCompressedParam.timestamp.desc()).first()

                if record and record.compressed_payload:
                    decompressed = gzip.decompress(record.compressed_payload).decode('utf-8')
                    payload = json.loads(decompressed)
                    return {
                        'topic': topic_name,
                        'payload': payload,
                        'timestamp': str(record.timestamp)
                    }
            except Exception as e:
                print(f"[规则评估] 从数据库获取压缩数据失败 topic={topic_name}: {e}")

        return None

    # 加载启用的规则配置（基础规则 + 设备规则）
    config_by_status_type = load_device_rules(db, device_code, DeviceStatusMonitorConfig)

    if not config_by_status_type:
        return {
            "device_code": device_code,
            "old_status": old_status,
            "new_status": old_status,
            "matched_statuses": [],
            "message": "无规则配置，状态未变更"
        }

    # 评估所有规则
    mutually_exclusive_statuses = ["processing", "stop", "fault_stop", "emergency stop", "mold_change", "maintain", "material_shortage"]
    coexistent_statuses = ["alarm"]

    matched_statuses = {}
    for status_type, configs in config_by_status_type.items():
        for config in configs:
            try:
                if _evaluate_config(config, get_latest_mqtt_message):
                    matched_statuses[status_type] = config
                    break
            except Exception as e:
                print(f"规则评估失败 (config_id={config.id}): {e}")

    # 曲线匹配评估
    all_configs = []
    for configs in config_by_status_type.values():
        all_configs.extend(configs)
    configs_with_curves = [c for c in all_configs if c.curve_id is not None]
    if configs_with_curves:
        # 获取事件消息用于曲线起始时间
        event_topics = [tc for tc in topic_configs_list if tc.topic_type == 'event']
        latest_event_msg = None
        for et in event_topics:
            msg = get_latest_mqtt_message(et.topic_name)
            if msg:
                if not latest_event_msg or msg.get('timestamp', '') > latest_event_msg.get('timestamp', ''):
                    latest_event_msg = msg

        curve_start_time = None
        if latest_event_msg:
            msg_timestamp = latest_event_msg.get('timestamp', '')
            if msg_timestamp:
                try:
                    curve_start_time = datetime.fromisoformat(msg_timestamp)
                except (ValueError, TypeError) as e:
                    print(f"[曲线同步] 解析event时间戳失败: {e}")

        db_curves = db.query(DBParamCurve).filter(
            and_(
                DBParamCurve.device_code == device_code,
                DBParamCurve.enabled == 1,
                DBParamCurve.id.in_([c.curve_id for c in configs_with_curves])
            )
        ).all()
        curve_map = {curve.id: curve for curve in db_curves}

        for config in configs_with_curves:
            curve = curve_map.get(config.curve_id)
            if not curve:
                continue
            try:
                if curve_start_time:
                    result = evaluate_curve_config(
                        config=config,
                        curve=curve,
                        mqtt_buffer=mqtt_buffer,
                        curve_start_time=curve_start_time,
                        field_path=config.field_path
                    )
                else:
                    result = evaluate_curve_config_from_buffer(
                        config=config,
                        curve=curve,
                        mqtt_buffer=mqtt_buffer,
                        field_path=config.field_path
                    )
                if result['matched']:
                    st = result['status_type']
                    if st not in matched_statuses or config.priority < matched_statuses[st].priority:
                        matched_statuses[st] = config
                        print(f"[曲线同步] 状态 {st} 匹配成功 score={result['score']}%")
            except Exception as e:
                print(f"[曲线同步] 评估失败 config_id={config.id}: {e}")

    # 默认状态映射：status_type → Device.status
    default_status_mapping = {
        "processing": "scheduled processing",
        "stop": "scheduled outage",
        "fault_stop": "maintenance",
        "emergency stop": "emergency stop",
        "mold_change": "mold_change",
        "maintain": "maintain",
    }

    # 解析互斥状态（按优先级），从规则配置中读取目标设备状态
    matched_mutually_exclusive = []
    for status_type in mutually_exclusive_statuses:
        if status_type in matched_statuses:
            config = matched_statuses[status_type]
            matched_mutually_exclusive.append({
                "status_type": status_type,
                "priority": config.priority,
                "device_status": getattr(config, 'device_status', None)
            })

    final_status_type = None
    new_status_parts = []
    if matched_mutually_exclusive:
        matched_mutually_exclusive.sort(key=lambda x: x["priority"])
        final = matched_mutually_exclusive[0]
        final_status_type = final["status_type"]
        if final["device_status"]:
            new_status_parts.append(final["device_status"])
        elif final_status_type in default_status_mapping:
            new_status_parts.append(default_status_mapping[final_status_type])
        else:
            new_status_parts.append(final_status_type)

    # 共存状态 alarm 追加到状态中
    if "alarm" in matched_statuses:
        new_status_parts.append("alarm")

    new_status = ','.join(new_status_parts) if new_status_parts else 'unknown'

    # 更新设备状态
    if new_status != old_status:
        device.status = new_status
        device.updated_at = datetime.now()
        db.commit()

    return {
        "device_code": device_code,
        "old_status": old_status,
        "new_status": new_status,
        "matched_statuses": list(matched_statuses.keys()),
        "final_mutually_exclusive": final_status_type,
        "message": f"状态已从 {old_status} 更新为 {new_status}" if new_status != old_status else "状态未变更"
    }


@router.post("/{device_code}/data-source/unsubscribe")
async def unsubscribe_device_realtime_data(
    device_code: str,
    topic_name: str = Query(..., description="Topic名称"),
    db: Session = Depends(get_db)
):
    """
    取消订阅设备实时数据

    Args:
        device_code: 设备编号
        topic_name: Topic名称

    Returns:
        取消订阅结果
    """
    device = db.query(Device).filter(Device.device_code == device_code).first()
    if not device:
        raise HTTPException(status_code=404, detail="设备不存在")

    collector = get_device_data_collector()
    success = collector.unsubscribe_realtime(device_code=device_code, topic_name=topic_name)

    if success:
        return {
            "message": "取消订阅成功",
            "device_code": device_code,
            "topic_name": topic_name,
            "status": "unsubscribed"
        }
    else:
        raise HTTPException(status_code=500, detail="取消订阅失败")


@router.get("/{device_code}/events")
async def get_device_events(
    device_code: str,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    start_time: Optional[int] = None,
    end_time: Optional[int] = None,
    db: Session = Depends(get_db),
):
    """获取设备关联的加工事件"""
    device = db.query(Device).filter(Device.device_code == device_code).first()
    if not device:
        raise HTTPException(status_code=404, detail="设备不存在")

    query = db.query(EventData).filter(EventData.machine_id == device_code)

    if start_time:
        query = query.filter(EventData.start_time >= start_time)
    if end_time:
        query = query.filter(EventData.start_time <= end_time)

    total = query.count()
    items = query.order_by(EventData.start_time.desc()).offset(
        (page - 1) * page_size
    ).limit(page_size).all()

    return {
        "total": total,
        "items": [
            {
                "id": e.id,
                "event_uid": e.event_uid,
                "start_code": e.start_code,
                "start_time": e.start_time,
                "end_time": e.end_time,
                "duringtime": e.duringtime,
                "machine_id": e.machine_id,
                "operator_name": e.operator_name,
                "process_no": e.process_no,
                "line_code": e.line_code,
                "created_at": str(e.created_at) if e.created_at else None,
            }
            for e in items
        ],
        "device": {
            "device_code": device.device_code,
            "device_name": device.device_name,
        }
    }
